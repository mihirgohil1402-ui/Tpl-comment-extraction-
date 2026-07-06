# app.py - TPL Comment Extractor (multi-provider, resilient, token-aware)
#
# Setup:
#   pip install streamlit aiohttp openpyxl pdfplumber
#   (optional OCR) pip install pytesseract pillow  + Tesseract binary
#   streamlit run app.py
#
# Single-file Streamlit app. Upload submittal ZIPs (or standalone submittal
# PDFs) -> reviewer comments are extracted by an LLM -> written to the exact
# TPL_Comments.xlsx format.

import streamlit as st
import zipfile
import io
import pdfplumber
import asyncio
import aiohttp
import json
import re
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import tempfile
import os
import gc

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# pdfminer (bundled with pdfplumber) decodes PDF text strings (UTF-16BE /
# PDFDocEncoding). Used when reading annotation /Contents raw bytes.
try:
    from pdfminer.utils import decode_text as _decode_pdf_text
except Exception:
    def _decode_pdf_text(b):
        if b[:2] == b"\xfe\xff":
            return b[2:].decode("utf-16-be", "replace")
        return b.decode("latin-1", "replace")

st.set_page_config(page_title="TPL Comment Extractor", layout="wide")

# ============================================================================
# CONFIG
# ============================================================================
CONCURRENCY = 2

# --- Request-sizing / retry knobs (safe defaults for free tiers) ------------
CHAR_BUDGET       = 24000   # max chars of PDF text per request (pre-split)
MODEL_CHAR_LIMIT  = 48000   # hard ceiling; above this we split into chunks
MAX_RETRIES       = 4       # attempts per provider before falling back
BACKOFF_BASE      = 2.0     # exponential backoff base (seconds): 2,4,8,...
BACKOFF_CAP       = 30.0    # never wait more than this between retries
REQUEST_TIMEOUT   = 90      # seconds per HTTP request
CHARS_PER_TOKEN   = 4.0     # rough token estimate (≈4 chars/token English)
MAX_CHUNKS_PER_DOC = 12     # emergency cap on requests per document; only a
                            # pathological doc ever hits this (12 x 24k chars)

# ----------------------------------------------------------------------------
# SUPPORTED APIS — add a new LLM service by adding an entry here.
#   url           : the chat/completions endpoint
#   format        : "openai" (Groq, OpenAI, Mistral, DeepSeek, Gemini-compat,
#                   Together...) or "anthropic" (Claude)
#   models        : list of model IDs (first is the dropdown default)
#   extra_headers : (optional) dict of extra headers, e.g. Claude's version
# ----------------------------------------------------------------------------
SUPPORTED_APIS = {
    "groq": {
        "url": "https://api.groq.com/openai/v1/chat/completions",
        "format": "openai",
        "models": ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"],
    },
    "openai": {
        "url": "https://api.openai.com/v1/chat/completions",
        "format": "openai",
        "models": ["gpt-4o", "gpt-4-turbo", "gpt-4o-mini"],
    },
    "claude": {
        "url": "https://api.anthropic.com/v1/messages",
        "format": "anthropic",
        "models": ["claude-opus-4-8", "claude-sonnet-4-6", "claude-haiku-4-5"],
        "extra_headers": {"anthropic-version": "2023-06-01"},
    },
    "gemini": {
        "url": "https://generativelanguage.googleapis.com/v1beta/openai/chat/completions",
        "format": "openai",
        "models": ["gemini-2.5-flash", "gemini-2.5-flash-lite", "gemini-2.5-pro"],
    },
    "kimi": {
        "url": "https://api.moonshot.ai/v1/chat/completions",
        "format": "openai",
        "models": ["kimi-k2.6", "kimi-k2.5", "kimi-latest"],
    },
    "mistral": {
        "url": "https://api.mistral.ai/v1/chat/completions",
        "format": "openai",
        "models": ["mistral-large-latest", "mistral-small-latest"],
    },
    "deepseek": {
        "url": "https://api.deepseek.com/v1/chat/completions",
        "format": "openai",
        "models": ["deepseek-chat"],
    },
    "together": {
        "url": "https://api.together.ai/v1/chat/completions",
        "format": "openai",
        "models": ["meta-llama/Llama-3-70b-chat-hf", "openai/gpt-oss-20b", "mistralai/Mistral-7B-Instruct-v0.1"],
    },
}

# Which HTTP statuses are worth retrying (transient). 429 = rate limit;
# 5xx = server hiccup. 4xx (except 429) are permanent -> don't waste retries.
RETRYABLE_STATUS = {429, 500, 502, 503, 504}


# ============================================================================
# PROVIDER POOL — key-level scheduling
# ============================================================================
# Instead of one key per provider with provider-level retries, we keep a POOL
# of keys. Each key tracks its own stats + cooldown. The scheduler hands out
# the next available key; a key that hits 429 is put on cooldown (not retried
# repeatedly) so the very next request uses a DIFFERENT key. Only when every
# key across every provider is cooling down do we wait for the soonest one.
#
# Scheduling priority (as specified):
#   available Gemini keys -> available Groq keys -> available Kimi keys ->
#   cooled-down Gemini keys -> cooled-down Groq keys -> cooled-down Kimi keys
# "available" = no active cooldown. Within a provider, keys are tried in the
# order the user entered them (round-robin so load spreads across a provider).

# Provider priority order for scheduling. Providers not listed here still work
# (they fall in after these, in dict order), keeping the abstraction general.
PROVIDER_PRIORITY = ["gemini", "groq", "kimi"]


class ApiKey:
    """One API key with its own independent stats and cooldown."""
    def __init__(self, provider, key, label):
        self.provider = provider
        self.key = key
        self.label = label            # e.g. "gemini#1" for analytics
        # per-key stats (requirement 8)
        self.retry_count = 0          # times this key was put on cooldown
        self.cooldown_until = 0.0     # epoch seconds; 0 = available now
        self.last_request_time = 0.0
        self.total_requests = 0
        self.total_failures = 0
        self.total_429 = 0
        self._429_streak = 0          # consecutive 429s -> grows backoff
        self.dead = False             # permanent auth failure (401/403)

    def available(self, now):
        return now >= self.cooldown_until

    def cooldown_remaining(self, now):
        return max(0.0, self.cooldown_until - now)

    def mark_429(self, now, retry_after=None):
        """Put this key on exponential-backoff cooldown. Doesn't retry it now.
        Honours the server's Retry-After when it is longer than our own
        backoff, so we never knock on a door the server told us is closed."""
        self.total_429 += 1
        self.total_failures += 1
        self.retry_count += 1
        self._429_streak += 1
        wait = min(BACKOFF_BASE ** self._429_streak, BACKOFF_CAP)
        if retry_after:
            wait = max(wait, min(float(retry_after), 120.0))
        self.cooldown_until = now + wait

    def mark_other_failure(self, now):
        """Non-429 failure (5xx/timeout): short cooldown, smaller penalty."""
        self.total_failures += 1
        self.retry_count += 1
        wait = min(BACKOFF_BASE ** 1, BACKOFF_CAP)
        self.cooldown_until = now + wait

    def mark_dead(self):
        """Permanent auth failure (401/403): never schedule this key again."""
        self.dead = True
        self.total_failures += 1
        self.retry_count += 1

    def mark_success(self):
        self._429_streak = 0          # reset backoff growth on success

    def stats_row(self):
        return {
            "key": self.label,
            "provider": self.provider,
            "total_requests": self.total_requests,
            "total_failures": self.total_failures,
            "total_429": self.total_429,
            "retry_count": self.retry_count,
        }


class KeyPool:
    """Holds every configured key and hands out the next one to use.

    provider_key_lists: {provider: [key1, key2, ...]} in the order entered.
    Backward compatible: a single gemini key + single groq key behaves exactly
    like the old primary+fallback flow (gemini used until it 429s, then groq)."""

    def __init__(self, provider_key_lists):
        self.keys = []
        for provider, key_list in provider_key_lists.items():
            for i, k in enumerate(key_list, 1):
                if k and k.strip():
                    self.keys.append(ApiKey(provider, k.strip(), f"{provider}#{i}"))
        # round-robin cursor per provider so we spread load across a provider's
        # keys instead of always hammering its first key.
        self._rr = {}

    def has_any(self):
        return len(self.keys) > 0

    def has_usable(self):
        return any(not k.dead for k in self.keys)

    def _provider_rank(self, provider):
        if provider in PROVIDER_PRIORITY:
            return PROVIDER_PRIORITY.index(provider)
        return len(PROVIDER_PRIORITY)  # unlisted providers come after

    def acquire(self, now):
        """Return (api_key, wait_seconds).
        - If a key is available now: (key, 0.0).
        - If all keys are cooling down: (soonest_key, seconds_to_wait) so the
          caller can sleep then use it (requirement 6: retry oldest exhausted).
        - If pool is empty: (None, 0.0)."""
        # Dead keys (401/403) are never scheduled again.
        alive = [k for k in self.keys if not k.dead]
        if not alive:
            return None, 0.0

        # 1) Prefer any AVAILABLE key, ordered by provider priority then a
        #    round-robin tiebreak within the pool.
        available = [k for k in alive if k.available(now)]
        if available:
            def sort_key(k):
                rr = self._rr.get(k.provider, 0)
                idx = self.keys.index(k)
                return (self._provider_rank(k.provider),
                        (idx - rr) % len(self.keys))
            available.sort(key=sort_key)
            chosen = available[0]
            self._rr[chosen.provider] = self.keys.index(chosen) + 1
            return chosen, 0.0

        # 2) Nothing available -> pick the key whose cooldown expires SOONEST.
        #    Tie-break by provider priority. This implements "retry the oldest
        #    exhausted key after backoff", preferring Gemini then Groq.
        soonest = min(
            alive,
            key=lambda k: (k.cooldown_until, self._provider_rank(k.provider))
        )
        return soonest, soonest.cooldown_remaining(now)

    def stats(self):
        return [k.stats_row() for k in self.keys]


# ============================================================================
# ANALYTICS — one row per HTTP request, aggregated into a summary at the end
# ============================================================================
class Analytics:
    """Thread/async-safe-enough for our single-loop use. Records every request."""
    def __init__(self):
        self.rows = []
        self.request_counter = 0

    def next_request_no(self):
        self.request_counter += 1
        return self.request_counter

    def record(self, **row):
        self.rows.append(row)

    def summary(self, total_zips, ok_zips, fail_zips, wall_seconds):
        calls      = len(self.rows)
        in_tokens  = sum(r.get("est_input_tokens", 0)  for r in self.rows)
        out_tokens = sum(r.get("est_output_tokens", 0) for r in self.rows)
        chars_sent = [r.get("chars_sent", 0) for r in self.rows]
        avg_size   = (sum(chars_sent) / len(chars_sent)) if chars_sent else 0
        largest    = max(chars_sent) if chars_sent else 0
        err_429    = sum(1 for r in self.rows if r.get("status") == 429)
        retries    = sum(r.get("retries", 0) for r in self.rows)
        fallbacks  = sum(1 for r in self.rows if r.get("fallback_used"))
        return {
            "Total ZIPs": total_zips,
            "Successful ZIPs": ok_zips,
            "Failed ZIPs": fail_zips,
            "Total API Calls": calls,
            "Total Estimated Input Tokens": in_tokens,
            "Total Estimated Output Tokens": out_tokens,
            "Average Request Size (chars)": round(avg_size),
            "Largest Request (chars)": largest,
            "429 Errors": err_429,
            "Retries": retries,
            "Fallbacks": fallbacks,
            "Total Processing Time (s)": round(wall_seconds, 1),
        }

    def per_key_table(self, key_pool):
        """Per-key analytics (requirement 9). Merges pool stats with request
        rows so each key shows its own request count, failures, 429s, latency."""
        # aggregate latency per key label from the request rows
        by_key = {}
        for r in self.rows:
            lbl = r.get("key_label")
            if lbl is None:
                continue
            d = by_key.setdefault(lbl, {"latency_sum": 0.0, "calls": 0})
            d["latency_sum"] += r.get("latency", 0.0)
            d["calls"] += 1
        table = []
        for ks in key_pool.stats():
            lbl = ks["key"]
            agg = by_key.get(lbl, {"latency_sum": 0.0, "calls": 0})
            avg_lat = (agg["latency_sum"] / agg["calls"]) if agg["calls"] else 0.0
            table.append({
                "Key": lbl,
                "Provider": ks["provider"],
                "Total Requests": ks["total_requests"],
                "Total Failures": ks["total_failures"],
                "Total 429": ks["total_429"],
                "Retry Count": ks["retry_count"],
                "Avg Latency (s)": round(avg_lat, 2),
            })
        return table


def est_tokens(text):
    """Rough token estimate. Good enough for staying under provider limits."""
    return int(len(text) / CHARS_PER_TOKEN) + 1


# ============================================================================
# DEBUG MODE — purely diagnostic pipeline tracing (never changes extraction)
# ============================================================================
# When "Enable Debug Mode" is ticked in the UI, every submittal writes a full
# pipeline trace to debug_logs/<SUBMITTAL>/ as numbered text files: which PDFs
# were selected and why, per-page extracted text, every annotation seen, the
# exact prompts sent, the raw LLM responses, the cleaned comments with a
# best-effort source attribution (page + annotation/page-text), the Excel rows
# and a summary. Every hook below is a no-op when trace is None, so disabled
# mode behaves exactly as before.

DEBUG_DIR = "debug_logs"


class DebugTrace:
    """Collects one submittal's pipeline artifacts; save() writes the files."""

    def __init__(self, sub_id):
        self.sub_id = sub_id
        self.selected = []     # [{pdf, reason}]
        self.skipped = []      # pdf names not selected as comment-bearing
        self.notes = []        # free-form diagnostics (read failures etc.)
        self.pages = []        # [{pdf, page, text, ocr_used, annots, included}]
        self.annotations = []  # [{pdf, page, subtype, text, kept, note}]
        self.prompts = []      # exact prompt string per chunk
        self.responses = []    # [{chunk, content, error, parse_error}]
        self.chunks = 0

    def note_selected(self, pdf, reason):
        self.selected.append({"pdf": pdf, "reason": reason})

    def _attribute(self, comment):
        """Best-effort source attribution for a CLEANED comment: find the page
        whose annotation (checked first — most comments are markups) or text
        layer contains it. Tries the full normalised comment, then shrinking
        prefixes, because the cleaning pipeline may have split or trimmed what
        the LLM returned. Diagnostic only — never affects extraction."""
        c = re.sub(r'\W+', ' ', comment).strip().lower()
        for key in (c, c[:80], c[:40]):
            if len(key) < 15:
                continue
            for a in self.annotations:
                if not a.get("kept"):
                    continue
                an = re.sub(r'\W+', ' ', a.get("text", "")).strip().lower()
                if not an:
                    continue
                if key in an or (len(an) >= 15 and an in c):
                    return a["pdf"], a["page"], "annotation"
            for p in self.pages:
                pn = re.sub(r'\W+', ' ', p["text"]).strip().lower()
                if key in pn:
                    return p["pdf"], p["page"], "page text"
        return None, None, ("not located verbatim (LLM likely paraphrased "
                            "or merged source text)")

    def save(self, base_dir, result, analytics):
        safe = re.sub(r'[^A-Za-z0-9._-]+', '_', self.sub_id) or "submittal"
        d = os.path.join(base_dir, safe)
        os.makedirs(d, exist_ok=True)

        def w(fname, text):
            with open(os.path.join(d, fname), "w", encoding="utf-8",
                      errors="replace") as f:
                f.write(text if text.endswith("\n") else text + "\n")

        # -- 01: which PDFs were selected as comment-bearing, and why --------
        lines = [f"Submittal: {self.sub_id}", ""]
        if self.selected:
            lines.append("Selected comment-bearing PDF(s):")
            for s in self.selected:
                lines.append(f"  - {s['pdf']}")
                lines.append(f"    reason: {s['reason']}")
        else:
            lines.append("No comment-bearing PDF selected -> submittal "
                         "treated as having no reviewer comments.")
        if self.skipped:
            lines += ["", "Skipped PDF(s) (no comment keyword in filename, "
                          "no sentence-like reviewer annotations):"]
            lines += [f"  - {n}" for n in self.skipped]
        if self.notes:
            lines += ["", "Notes:"] + [f"  - {n}" for n in self.notes]
        w("01_selected_pdf.txt", "\n".join(lines))

        # -- 02: raw per-page extracted text ---------------------------------
        parts = ["Raw per-page text as extracted (after OCR fallback, BEFORE "
                 "boilerplate dedupe / chunking).", ""]
        for p in self.pages:
            parts.append(f"{'=' * 22} {p['pdf']} — page {p['page']} {'=' * 22}")
            parts.append(f"[OCR used: {'yes' if p['ocr_used'] else 'no'}] "
                         f"[included in pipeline: "
                         f"{'yes' if p['included'] else 'no (empty page)'}]")
            parts.append(p["text"].strip() or "<no extractable text>")
            parts.append("")
        if not self.pages:
            parts.append("No pages extracted.")
        w("02_extracted_text.txt", "\n".join(parts))

        # -- 03: every annotation seen, with page + subtype ------------------
        parts = []
        for a in self.annotations:
            parts.append(f"{a['pdf']} — page {a['page']}")
            parts.append(f"  subtype: {a['subtype'] or '(none)'}")
            kept = "yes" if a["kept"] else f"no ({a['note']})"
            parts.append(f"  kept:    {kept}")
            parts.append(f"  text:    {a['text'] or '<empty>'}")
            parts.append("")
        w("03_annotations.txt",
          "\n".join(parts) if self.annotations else "No annotations found.")

        # -- 04: exact prompt(s) sent to the LLM -----------------------------
        parts = []
        for i, pr in enumerate(self.prompts, 1):
            parts.append(f"{'=' * 20} prompt — chunk {i}/{len(self.prompts)} "
                         f"{'=' * 20}")
            parts.append(pr)
            parts.append("")
        w("04_prompt.txt",
          "\n".join(parts) if self.prompts
          else "No prompt sent (no text or no comment-bearing PDF).")

        # -- 05: raw LLM response(s) before parsing --------------------------
        parts = []
        for r in self.responses:
            parts.append(f"{'=' * 20} response — chunk {r['chunk']} {'=' * 20}")
            if r["error"]:
                parts.append(f"[request error: {r['error']}]")
            parts.append(r["content"] if r["content"] is not None
                         else "<no content>")
            if r.get("parse_error"):
                parts.append(f"[parse error: {r['parse_error']}]")
            parts.append("")
        w("05_llm_response.txt",
          "\n".join(parts) if self.responses else "No LLM call was made.")

        # -- 06: cleaned comments with source page attribution ---------------
        parts = []
        for i, c in enumerate(result.get("comments", []), 1):
            pdf, page, src = self._attribute(c)
            parts.append(f"--- Comment {i} ---")
            parts.append(f"Page {page}  ({pdf})" if page is not None
                         else "Page: not located")
            parts += ["", "Comment:", c, "", "Source:", src, ""]
        w("06_cleaned_comments.txt",
          "\n".join(parts) if result.get("comments")
          else "No comments extracted."
             + (f"\nError: {result['error']}" if result.get("error") else ""))

        # -- 07: the Excel rows this submittal produces ----------------------
        comments = result.get("comments", [])
        unread = result.get("unreadable", 0)
        note = (f"{unread} page(s) could not be read as text (possible "
                f"scanned comments) — check PDF manually") if unread else ""
        rows = ["Columns: Sr no. | Submittal | Document | Costumer Comments "
                "| Xylem Remarks | Review Note",
                "(Sr no. is assigned globally when the workbook is built; "
                "shown as '-')", ""]
        if comments:
            for k, cm in enumerate(comments, 1):
                rows.append(" | ".join([
                    "-" if k == 1 else "",
                    self.sub_id if k == 1 else "",
                    (result.get("document", "") or "") if k == 1 else "",
                    f"{k}. {cm}",
                    "",
                    note if k == 1 else ""]))
        else:
            rows.append(" | ".join(["-", self.sub_id,
                                    result.get("document", "") or "",
                                    "", "Comment not Received", note]))
        w("07_excel_rows.txt", "\n".join(rows))

        # -- 08: summary ------------------------------------------------------
        my = [r for r in analytics.rows if r.get("submittal") == self.sub_id]
        sizes = [r.get("chars_sent", 0) for r in my]
        lats = [r.get("latency", 0.0) for r in my]
        failed = sum(1 for r in my if r.get("status") != 200)
        lines = [
            f"Submittal:          {self.sub_id}",
            f"Document:           {result.get('document', '')}",
            f"Total pages:        {len(self.pages)}",
            f"OCR pages:          "
            f"{sum(1 for p in self.pages if p['ocr_used'])}",
            f"Annotation pages:   "
            f"{sum(1 for p in self.pages if p['annots'])}",
            f"Unreadable pages:   {unread}",
            f"Chunks:             {self.chunks}",
            f"API calls:          {len(my)}",
            f"Provider(s) used:   "
            f"{', '.join(sorted({r['provider'] for r in my})) or '-'}",
            f"Model(s):           "
            f"{', '.join(sorted({r['model'] for r in my})) or '-'}",
            f"Key(s):             "
            f"{', '.join(sorted({r['key_label'] for r in my})) or '-'}",
            f"Request size:       avg {round(sum(sizes) / len(sizes)) if sizes else 0} chars, "
            f"max {max(sizes) if sizes else 0} chars",
            f"Latency:            total {round(sum(lats), 2)}s, "
            f"avg {round(sum(lats) / len(lats), 2) if lats else 0}s",
            f"Retries (failed attempts): {failed}",
            f"429 responses:      "
            f"{sum(1 for r in my if r.get('status') == 429)}",
            f"Comments extracted: {len(comments)}",
            f"Error:              {result.get('error') or '-'}",
        ]
        w("08_summary.txt", "\n".join(lines))


# ============================================================================
# PDF / ZIP HELPERS
# ============================================================================
def submittal_id_from_name(file_name):
    """SUB1715-_ Rev No_ 0.zip  ->  SUB1715   (works for .pdf uploads too)"""
    m = re.search(r'(SUB\d+)', file_name, re.I)
    if m:
        return m.group(1).upper()
    return re.sub(r'\.(zip|pdf)$', '', file_name, flags=re.I)


def _pdf_items_from_upload(file_name, file_bytes):
    """Normalise an upload into [(pdf_name, pdf_bytes), ...].
    A .pdf upload is a one-item list; a .zip contributes every PDF inside."""
    if file_name.lower().endswith('.pdf'):
        return [(file_name, file_bytes)]
    items = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zf:
        for n in zf.namelist():
            if n.lower().endswith('.pdf'):
                items.append((n, zf.read(n)))
    return items


def _read_page_annotations(page, collect=None):
    """Reviewer markups saved as live annotation objects (FreeText callouts,
    notes) never appear in extract_text() — their text lives in the
    annotation's /Contents. Read it directly. Link/Popup/Widget annotations
    are navigation chrome, not comments; stamps/shapes without text are
    skipped automatically because their /Contents is empty.

    `collect` (Debug Mode only): when a list is supplied, every annotation
    object seen is appended as {subtype, text, kept, note}, including the
    skipped ones. Never changes what is returned."""
    texts = []
    try:
        annots = page.annots or []
    except Exception:
        return texts
    for a in annots:
        try:
            data = a.get("data") or {}
            sub = data.get("Subtype")
            sub_name = getattr(sub, "name", "") or (str(sub) if sub else "")
            if sub_name in ("Link", "Popup", "Widget"):
                if collect is not None:
                    collect.append({"subtype": sub_name, "text": "",
                                    "kept": False,
                                    "note": "navigation chrome — skipped"})
                continue
            raw = a.get("contents")
            if raw is None:
                raw = data.get("Contents")
            if isinstance(raw, bytes):
                try:
                    raw = _decode_pdf_text(raw)
                except Exception:
                    raw = raw.decode("latin-1", "replace")
            txt = re.sub(r'\s+', ' ', str(raw)).strip() if raw else ""
            if txt:
                texts.append(txt)
                if collect is not None:
                    collect.append({"subtype": sub_name, "text": txt,
                                    "kept": True, "note": ""})
            elif collect is not None:
                collect.append({"subtype": sub_name, "text": "",
                                "kept": False,
                                "note": "empty /Contents — skipped"})
        except Exception:
            continue
    return texts


# Filename keywords that mark a PDF as comment-bearing. Broader than just
# annotated/response so common conventions are caught by name alone.
COMMENT_NAME_KEYWORDS = ('annotated', 'response', 'markup', 'marked',
                         'comment', 'reply')


def _name_says_comments(pdf_name):
    nl = pdf_name.lower()
    return any(k in nl for k in COMMENT_NAME_KEYWORDS)


def _looks_like_reviewer_note(txt):
    """Sentence-like annotation text = reviewer comment. CAD drawings also
    carry FreeText annotations, but those are short ALL-CAPS fragments (part
    numbers, materials, 'ALL DIMENSION ARE IN MM') — never sentences with
    lowercase words."""
    return (len(txt) >= 30 and len(txt.split()) >= 5
            and re.search(r'[a-z]', txt) is not None)


def _has_reviewer_annotations(raw):
    """Does this PDF carry at least one sentence-like markup annotation?
    Used to catch comment-bearing PDFs whose filename doesn't say so."""
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                for t in _read_page_annotations(page):
                    if _looks_like_reviewer_note(t):
                        return True
    except Exception:
        pass
    return False


def _vendor_reference_text(pdf_items):
    """Normalised text of every UNMARKED base twin PDF in the upload.

    Vendor GA drawings carry their own notes blocks ('THE DRAWING SHALL BE
    USED FOR REFERENCE ONLY...', 'SKID SHOULD BE PROPERLY GROUTED.') whose
    imperative wording weaker models mistake for reviewer comments. Those
    lines exist verbatim in the unannotated base PDF shipped in the same
    ZIP; reviewer additions by definition do not. Extracted comments that
    are textually contained in the twin are therefore vendor content and
    can be removed deterministically AFTER the LLM step — the model input
    stays byte-identical, so recall of genuine comments is unaffected.

    Returns a single space-padded normalised string ('' when no twin)."""
    def canon(pdf_name):
        # names are equal modulo the 'annotated' token and trailing
        # separators before .pdf (some archives truncate long names, leaving
        # e.g. '...A-B-C-D-_annotated.pdf' next to '...A-B-C-D-.pdf')
        n = pdf_name.lower()
        n = re.sub(r'[\s_-]*annotated', '', n)
        return re.sub(r'[\s_-]+(?=\.pdf$)', '', n)

    chunks = []
    for name, _ in pdf_items:
        if 'annotated' not in name.lower():
            continue
        for n2, b2 in pdf_items:
            if n2 == name or 'annotated' in n2.lower() \
                    or canon(n2) != canon(name):
                continue
            try:
                with pdfplumber.open(io.BytesIO(b2)) as pdf:
                    for p in pdf.pages:
                        chunks.append(p.extract_text() or "")
            except Exception:
                pass
    if not chunks:
        return ""
    blob = re.sub(r'[^a-z0-9]+', ' ', " ".join(chunks).lower()).strip()
    return f" {blob} " if blob else ""


def _is_vendor_comment(comment, vendor_blob):
    """Is this cleaned LLM comment vendor content from the base document?

    Long comments are matched by 5-word shingles — the LLM stitches lines
    that PDF extraction wrapped/interleaved (drawing tables run through the
    notes column), so exact substring matching fails; >=80% of shingles
    found in the twin means the comment is vendor text. Calibrated on the
    corpus: vendor notes score 0.84-1.00, genuine reviewer comments score
    0.00, and a pathological extraction line mixing a vendor table row with
    a comment fragment scores 0.71 — safely below the threshold.

    Column-interleaved vendor text (two-column installation manuals) breaks
    shingle contiguity: pdfplumber merges columns per visual line, so the
    LLM's cleanly de-interleaved sentence no longer appears contiguously in
    the twin. Interleaving inserts words but never REORDERS them, so a
    second test matches the comment as an ordered subsequence of the twin
    text within a bounded window (4x the comment length).

    Short comments (2-4 words, e.g. a false 'VERY COARSE') must match
    exactly; single-word comments (a genuine 'Model' list item) are never
    vendor-classified."""
    if not vendor_blob:
        return False
    words = re.sub(r'[^a-z0-9]+', ' ', comment.lower()).split()
    if len(words) < 2:
        return False
    if len(words) < 5:
        return f" {' '.join(words)} " in vendor_blob
    shingles = [' '.join(words[i:i + 5]) for i in range(len(words) - 4)]
    hits = sum(1 for s in shingles if s in vendor_blob)
    if hits / len(shingles) >= 0.8:
        return True
    base = vendor_blob.split()
    max_span = 4 * len(words)
    for start in (i for i, w in enumerate(base) if w == words[0]):
        i, j = start, 0
        while i < len(base) and j < len(words) and i - start < max_span:
            if base[i] == words[j]:
                j += 1
            i += 1
        if j == len(words):
            return True
    return False


def extract_pdf_text(pdf_items, treat_all_as_comments=False, trace=None):
    """Pull text from the COMMENT-bearing PDFs using pdfplumber.

    Reviewer comments live only in an annotated markup PDF or a reviewer
    'Response' PDF. The plain datasheet and the submittal lead sheet contain
    NO reviewer comments — reading them makes the LLM invent "comments" out
    of spec-sheet text (false positives). Selection is therefore:
      - every PDF whose NAME contains a comment keyword (annotated, response,
        markup, comment, reply, ...), PLUS
      - every PDF whose CONTENT carries sentence-like markup annotations
        (a Response PDF can arrive with any filename — selecting on name
        alone made the ZIP path miss files the direct-PDF path processed).
      - If neither finds anything, the submittal has no reviewer comments.
      - treat_all_as_comments=True (a directly-uploaded PDF, which has no
        sibling files to compare against) reads every given PDF.

    Comments stored as live annotation objects (FreeText callouts in
    'Response' PDFs) are read via /Contents in page order; annotation text
    that is already flattened into the page text is not duplicated.

    Falls back to OCR if a page has images but no extractable text.

    Returns (pages_text, pages_annots, unreadable_pages, had_comment_file);
    pages_text[i] and pages_annots[i] describe the same page, in order.

    `trace` (Debug Mode only) records selection reasons, per-page text and
    annotations. All trace hooks are no-ops when trace is None.
    """
    pages_text = []
    pages_annots = []
    unreadable_pages = 0
    if treat_all_as_comments:
        comment_pdfs = list(pdf_items)
        if trace is not None:
            for n, _ in pdf_items:
                trace.note_selected(n, "direct PDF upload — no sibling files, "
                                       "treated as comment-bearing")
    else:
        chosen = {n for n, _ in pdf_items if _name_says_comments(n)}
        if trace is not None:
            for n in sorted(chosen):
                kw = next((k for k in COMMENT_NAME_KEYWORDS
                           if k in n.lower()), "?")
                trace.note_selected(
                    n, f"filename contains comment keyword '{kw}'")
        for n, b in pdf_items:
            if n not in chosen and _has_reviewer_annotations(b):
                chosen.add(n)
                if trace is not None:
                    trace.note_selected(n, "content carries sentence-like "
                                           "reviewer annotations")
        if trace is not None:
            trace.skipped = [n for n, _ in pdf_items if n not in chosen]
        # keep original archive order so comment order is stable
        comment_pdfs = [(n, b) for n, b in pdf_items if n in chosen]
        if not comment_pdfs:
            return [], [], 0, False
    for name, raw in comment_pdfs:
        try:
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page_no, page in enumerate(pdf.pages, 1):
                    t = page.extract_text() or ""
                    has_images = False
                    try:
                        has_images = len(page.images) > 0
                    except Exception:
                        pass

                    # OCR fallback: page has images but no real text
                    ocr_used = False
                    if len(t.strip()) < 50 and has_images and OCR_AVAILABLE:
                        try:
                            img = page.to_image()
                            ocr_text = pytesseract.image_to_string(img.original)
                            if len(ocr_text.strip()) > len(t.strip()):
                                t = ocr_text
                                ocr_used = True
                        except Exception:
                            pass

                    collect = [] if trace is not None else None
                    annot_texts = _read_page_annotations(page, collect)
                    # Avoid duplicates: skip annotation text that is already
                    # flattened into this page's text layer.
                    if annot_texts:
                        page_norm = re.sub(r'\s+', ' ', t).lower()
                        kept_texts = [x for x in annot_texts
                                      if x[:60].lower() not in page_norm]
                        if collect:
                            for entry in collect:
                                if entry["kept"] and \
                                        entry["text"] not in kept_texts:
                                    entry["kept"] = False
                                    entry["note"] = ("text already flattened "
                                                     "into page text — deduped")
                        annot_texts = kept_texts

                    if trace is not None:
                        trace.pages.append({
                            "pdf": name, "page": page_no, "text": t,
                            "ocr_used": ocr_used,
                            "annots": list(annot_texts),
                            "included": bool(t.strip() or annot_texts),
                        })
                        for entry in (collect or []):
                            trace.annotations.append(
                                {"pdf": name, "page": page_no, **entry})

                    if t.strip() or annot_texts:
                        pages_text.append(t)
                        pages_annots.append(annot_texts)
                    if len(t.strip()) < 50 and has_images:
                        unreadable_pages += 1
        except Exception as e:
            if trace is not None:
                trace.notes.append(f"failed to read {name}: {str(e)[:80]}")
    return pages_text, pages_annots, unreadable_pages, True


# --- Document name extraction ------------------------------------------------
# The Document column needs "{DOC_NO}_{Title}". Different PDFs in the ZIP label
# these differently, so we try several patterns and search EVERY pdf until we
# have both a number and a title (or the best partial we can find).

DOC_NO_RE = re.compile(r'\b(TPL-[A-Z0-9]+-\d+-[A-Z]+-[A-Z]+-\d+)\b', re.I)

# Fallback for documents that don't use the TPL-... scheme: take whatever
# identifier follows an explicit "Doc No / Document Number" label.
DOC_NO_FALLBACK_RE = re.compile(
    r'DOC(?:UMENT)?\.?\s*(?:NO|NUMBER)\.?\s*[:\-]?\s*'
    r'([A-Z][A-Z0-9][A-Z0-9\-/\.]{3,40}\d)',
    re.I
)

# A title value that starts with the document number ("<DOCNO>_<title>" — the
# lead sheet's Subject field is exactly this) gives both pieces at once.
DOC_NO_PREFIX_RE = re.compile(
    r'^(TPL-[A-Z0-9]+-\d+-[A-Z]+-[A-Z]+-\d+)[\s_:\-]+(.+)$', re.I | re.S)

# What ENDS a title value. Deliberately precise: 'DOC NO:', 'Rev 0',
# 'Sheet 1', 'Page 2' end a title — but the WORDS 'sheet'/'rev' inside a
# title ("Pump Technical Data sheet for CWTP...") must never truncate it
# (that truncation was why the Document column lost half its titles).
_TITLE_STOP = (
    r'DOC(?:UMENT)?\.?\s*(?:NO|NUMBER)\b'
    r'|Rev(?:ision)?\.?\s*(?:No)?\.?\s*[:.]?\s*\d'
    r'|Sheet\s*(?:No\b|\d)'
    r'|Page\s*(?:No\b|\d)'
    r'|Document\s*/\s*Dwg'
    r'|Location\s*:'
    r'|List\s+Attachments'
    r'|$'
)


def _title_rx(label):
    return re.compile(label + r'\s*[:\-]\s*(.+?)\s*(?:' + _TITLE_STOP + r')',
                      re.I | re.S)


# Ordered label list: (priority, compiled regex). All matches from every
# label on every PDF are collected; the best candidate wins (priority first,
# then completeness/length). A candidate carrying the DOCNO_ prefix gets a
# large boost — it is the exact target format.
TITLE_LABELS = [
    (4, _title_rx(r'(?:Document|Doc\.?|Drawing|Drg\.?)\s*Title')),
    (3, _title_rx(r'(?:Document|Doc\.?)\s*Name')),
    (3, _title_rx(r'Name\s*of\s*(?:the\s*)?Document')),
    (2, _title_rx(r'\bSubject')),
    (2, _title_rx(r'\bTitle')),
    (1, _title_rx(r'\bDescription')),
    (1, _title_rx(r'\bName')),
]


def _clean_title(raw):
    t = re.sub(r'\s+', ' ', raw).strip(" :-\t\r\n")
    t = re.sub(r'(\w)-\s+(\w)', r'\1-\2', t)   # rejoin hyphen-split words
    # Drop trailing submittal/project boilerplate that isn't part of the title.
    t = re.split(r'\b(?:Submittal\s*Number|Submittal\s*No|Project\b|Dholera)\b',
                 t, flags=re.I)[0].strip(" :-.")
    return t


def _title_candidates(full_text):
    """All labelled title candidates in one PDF's text.
    Returns a list of (priority, title, doc_no_or_empty)."""
    out = []
    for prio, rx in TITLE_LABELS:
        for m in rx.finditer(full_text):
            cand = _clean_title(m.group(1))
            if not (0 < len(cand) <= 160):
                continue
            pm = DOC_NO_PREFIX_RE.match(cand)
            if pm:
                title = _clean_title(pm.group(2))
                if 0 < len(title) <= 160:
                    out.append((prio + 10, title, pm.group(1)))
            else:
                out.append((prio, cand, ""))
    return out


def _layout_titles(full_text):
    """Label-free fallback: in these header blocks the title lines sit
    directly ABOVE the 'DOC NO' line (below a bare 'Document Title:' label
    on its own line). Collect up to 3 preceding non-label lines."""
    out = []
    lines = full_text.split("\n")
    for i, ln in enumerate(lines):
        if re.match(r'\s*DOC(?:UMENT)?\.?\s*(?:NO|NUMBER)\b', ln, re.I):
            grab = []
            for prev in reversed(lines[max(0, i - 3):i]):
                s = prev.strip()
                if not s or s.endswith(':'):
                    break
                grab.insert(0, s)
            cand = _clean_title(" ".join(grab))
            if 0 < len(cand) <= 160 and re.search(r'[a-z]', cand):
                out.append(cand)
    return out


def _title_from_filenames(pdf_items):
    """Last resort: attachment PDFs are named '001_<title>.pdf'."""
    for name, _ in pdf_items:
        base = name.rsplit('/', 1)[-1]
        base = re.sub(r'\.pdf$', '', base, flags=re.I)
        base = re.sub(r'_annotated$', '', base, flags=re.I)
        m = re.match(r'^\d{2,4}[_\- ]+(.{8,160})$', base)
        if m and re.search(r'[a-z]', m.group(1)):
            return _clean_title(m.group(1))
    return ""


def extract_doc_name_from_items(pdf_items):
    """Search EVERY PDF (from a ZIP or a single upload) for the document name.

    Strategy (all candidates are collected before choosing — never stop at
    the first match, and number/title may come from DIFFERENT PDFs):
      1. Labelled fields on the first pages of every PDF (Document Title,
         Drawing Title, Document Name, Subject, Title, Description, Name).
         A value shaped '<DOCNO>_<title>' (the lead sheet's Subject field)
         wins outright — it is the exact target format.
      2. Best labelled candidate by (priority, completeness/length).
      3. Header-layout fallback: title lines directly above the DOC NO line.
      4. Filename fallback: attachment PDFs named '001_<title>.pdf'.
      5. Only return a partial result if every strategy failed.
    """
    best_doc_no = ""
    candidates = []       # (priority, title)
    layout = []           # label-free fallback candidates
    for name, raw in pdf_items:
        try:
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                pages = []
                for p in pdf.pages[:3]:
                    t = p.extract_text() or ""
                    if t.strip():
                        pages.append(t)
            full = "\n".join(pages)
        except Exception:
            continue

        if not best_doc_no:
            m = DOC_NO_RE.search(full) or DOC_NO_FALLBACK_RE.search(full)
            if m:
                best_doc_no = m.group(1).strip()
        for prio, title, docno in _title_candidates(full):
            if docno and not best_doc_no:
                best_doc_no = docno
            candidates.append((prio, title))
        layout.extend(_layout_titles(full))

    best_title = ""
    if candidates:
        best_title = max(candidates, key=lambda c: (c[0], len(c[1])))[1]
    elif layout:
        best_title = max(layout, key=len)
    if not best_title:
        best_title = _title_from_filenames(pdf_items)

    if best_doc_no and best_title:
        return f"{best_doc_no}_{best_title}"
    return best_doc_no or best_title or ""


# ============================================================================
# TOKEN REDUCTION — remove repeated boilerplate WITHOUT touching unique content
# ============================================================================
# Reviewer comments are, by definition, the lines that vary between pages.
# Headers / footers / revision blocks / spec boilerplate repeat verbatim on
# many pages. We drop a line only when the SAME normalised line appears on
# multiple pages. Anything unique (i.e. every real comment) is always kept.

def _norm_line(line):
    """Normalise for repetition detection: collapse spaces, strip page numbers."""
    s = re.sub(r'\s+', ' ', line).strip()
    s = re.sub(r'\bpage\s*\d+(\s*of\s*\d+)?\b', '', s, flags=re.I)
    s = re.sub(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b', '', s)   # dates
    return s.lower().strip()


def dedupe_boilerplate(pages_text):
    """Drop REPEATS of lines that appear on multiple pages (headers/footers).

    Keeps:
      - every line that appears on only one page (unique content, incl. comments)
      - the FIRST occurrence of a repeated line. Repeated lines are almost
        always headers/footers, but a genuine comment stamped on two pages
        would previously vanish entirely — keeping one copy costs a few
        tokens and can never lose content (duplicate extracted comments are
        removed after the LLM step anyway).
    Safety: if a page would be emptied entirely, keep it as-is (avoid nuking a
    page whose comment happens to resemble a header)."""
    if len(pages_text) < 2:
        return pages_text  # nothing repeats across a single page

    # Count how many pages each normalised line appears on, and remember the
    # first page it was seen on (that copy is the one we keep).
    from collections import defaultdict
    page_count = defaultdict(int)
    first_page = {}
    per_page_lines = []
    for pi, t in enumerate(pages_text):
        lines = t.split("\n")
        per_page_lines.append(lines)
        seen_here = set()
        for ln in lines:
            n = _norm_line(ln)
            if n and n not in seen_here:
                seen_here.add(n)
                page_count[n] += 1
                if n not in first_page:
                    first_page[n] = pi

    # A line is boilerplate if it shows up on >= 2 pages AND is short-ish
    # (real multi-page comment paragraphs are rare; headers/footers are short).
    def is_boiler(nline):
        return page_count.get(nline, 0) >= 2 and len(nline) <= 120

    cleaned_pages = []
    for pi, lines in enumerate(per_page_lines):
        kept = []
        for ln in lines:
            n = _norm_line(ln)
            if n and is_boiler(n) and first_page.get(n) != pi:
                continue          # repeat copy -> drop; first copy is kept
            kept.append(ln)
        # Safety net: never let dedupe empty a page completely.
        if not any(l.strip() for l in kept):
            kept = lines
        cleaned_pages.append("\n".join(kept))
    return cleaned_pages


# Words/phrases that signal a page carries reviewer comments rather than
# pure spec-sheet boilerplate. Used by cap_chunks to rank chunks if a doc is
# pathologically long.
COMMENT_SIGNALS = re.compile(
    r'\b(shall|should|to be|required|ensure|provide|verify|comply|complian'
    r'|approved|vendor|deviation|clarification|as per|not acceptable'
    r'|to be furnished|to be considered|consult|review)\b',
    re.I
)


def cap_chunks(chunks, max_chunks=MAX_CHUNKS_PER_DOC):
    """Last-resort guard for pathological documents: if even chunking would
    mean an unreasonable number of API calls, keep the chunks with the most
    comment-signal words (in document order). Only here can content ever be
    dropped — normal documents always send everything."""
    if len(chunks) <= max_chunks:
        return chunks
    ranked = sorted(range(len(chunks)),
                    key=lambda i: -len(COMMENT_SIGNALS.findall(chunks[i])))
    keep = sorted(ranked[:max_chunks])
    return [chunks[i] for i in keep]


def split_oversized(text, limit=MODEL_CHAR_LIMIT):
    """If a request is still too big after prioritisation, split on page
    boundaries into <=limit chunks. Returns a list of chunk strings.
    Splitting on '\n' keeps whole lines (never cuts a comment mid-sentence)."""
    if len(text) <= limit:
        return [text]
    chunks, cur, cur_len = [], [], 0
    for line in text.split("\n"):
        add = len(line) + 1
        if cur_len + add > limit and cur:
            chunks.append("\n".join(cur))
            cur, cur_len = [], 0
        cur.append(line)
        cur_len += add
    if cur:
        chunks.append("\n".join(cur))
    return chunks


# ============================================================================
# COMMENT SPLITTER / CLEANERS  (unchanged behaviour)
# ============================================================================
def split_merged_comments(comments_list):
    """Split blocks like '1) aaa 2) bbb 3) ccc' (or '1. aaa 2. bbb') into
    separate comments. Only splits when a REAL numbered sequence is present:
    at least two markers, starting at 1 or 2, strictly increasing. A lone
    'n)' inside a sentence (clause references, '9-pin (SUB-D)', bore sizes)
    no longer splits a comment in half, and decimals like '2.4' never match
    because the marker requires whitespace after the separator."""
    marker = re.compile(r'(?:(?<=\s)|^)(\d{1,2})[\.\)]\s+')
    result = []
    for comment in comments_list:
        ms = list(marker.finditer(comment))
        nums = [int(m.group(1)) for m in ms]
        is_seq = (len(ms) >= 2 and nums[0] <= 2 and
                  all(b > a for a, b in zip(nums, nums[1:])))
        if not is_seq:
            result.append(comment.strip())
            continue
        parts = []
        pre = comment[:ms[0].start()].strip(' .')
        if pre:
            parts.append(pre)   # lead-in text; header filter handles it later
        for i, m in enumerate(ms):
            end = ms[i + 1].start() if i + 1 < len(ms) else len(comment)
            seg = comment[m.end():end].strip(' .')
            if seg:
                parts.append(seg)
        if parts:
            result.extend(parts)
        else:
            result.append(comment.strip())
    return result


HEADER_PATTERNS = re.compile(
    r'^(clarification|clarifications required|following points|following items'
    r'|please incorporate|please note the following|comments?:?$'
    r'|insert text here\.?$)',
    re.I
)


def drop_headers(comments_list):
    out = []
    for c in comments_list:
        if HEADER_PATTERNS.match(c.strip()):
            continue
        out.append(c)
    return out


LEADIN = re.compile(
    r'^(clarifications?\s+required[^:]*:\s*'
    r'|following\s+(?:points|items)[^:]*:\s*'
    r'|please\s+incorporate[^:]*:\s*'
    r'|please\s+note[^:]*:\s*)',
    re.I
)


def strip_leadin(text):
    return LEADIN.sub('', text).strip()


def dedupe_comments(comments):
    """Remove duplicate comments while preserving first-seen order. Chunked
    requests and markups repeated across pages can yield the same comment
    twice; the Excel should list it once."""
    seen = set()
    out = []
    for c in comments:
        k = re.sub(r'\W+', ' ', c).strip().lower()
        if k and k not in seen:
            seen.add(k)
            out.append(c)
    return out


# ============================================================================
# LLM CALL  (prompt wording unchanged)
# ============================================================================
PROMPT_TEMPLATE = """You are extracting reviewer comments from an engineering submittal review.

Below is the text of a submittal PDF. Reviewers have added comments/markups requesting clarifications or changes.

Extract EVERY distinct reviewer comment as its own separate item, in the order they appear in the text.

STRICT RULES:
- Each comment = ONE distinct instruction, requirement, or question.
- If several requirements are written together (numbered 1) 2) 3) OR just run together in a block), SPLIT them into separate items. Never combine two different requirements into one item.
- Do NOT include section headers or lead-ins such as "Clarifications required for the following points", "Following items to be mentioned", "Please incorporate following detail". Skip those; only output the actual items under them.
- Do NOT include document body text, specifications tables, titles, or boilerplate. Only the reviewer's added comments.
- Keep each comment's wording complete and faithful. Do not summarise or shorten.
- Output ONLY valid JSON, no commentary, in exactly this form:
{{"comments": ["first comment", "second comment", "third comment"]}}

SUBMITTAL TEXT:
{body}

JSON:"""


def _headers_and_payload(cfg, api_key, model, prompt):
    """Build (headers, payload) for a provider from its config."""
    fmt = cfg.get("format", "openai")
    headers = {"Content-Type": "application/json"}
    if fmt == "anthropic":
        headers["x-api-key"] = api_key
    else:
        headers["Authorization"] = f"Bearer {api_key}"
    for k, v in cfg.get("extra_headers", {}).items():
        headers[k] = v

    if fmt == "anthropic":
        payload = {
            "model": model,
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": prompt}],
        }
    else:
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 4096,
        }
    return headers, payload


def _extract_content(fmt, data):
    """Pull the text content out of a provider's JSON response."""
    if fmt == "anthropic":
        return data["content"][0]["text"]
    return data["choices"][0]["message"]["content"]


async def _single_call(session, cfg, api_key, model, prompt):
    """One HTTP attempt. Returns (content, error, status, retry_after).
    Exactly one of content/error is non-None. status is the HTTP code (or
    None on exception); retry_after is the server's Retry-After seconds on
    429/5xx when provided."""
    url = cfg["url"]
    fmt = cfg.get("format", "openai")
    headers, payload = _headers_and_payload(cfg, api_key, model, prompt)
    try:
        async with session.post(url, json=payload, headers=headers,
                                timeout=REQUEST_TIMEOUT) as resp:
            status = resp.status
            retry_after = None
            ra = resp.headers.get("Retry-After")
            if ra:
                try:
                    retry_after = float(ra)
                except ValueError:
                    retry_after = None
            if status == 200:
                data = await resp.json()
                try:
                    return _extract_content(fmt, data), None, status, None
                except (KeyError, IndexError, TypeError) as e:
                    return None, f"bad response shape: {str(e)[:60]}", status, None
            txt = await resp.text()
            return None, f"HTTP {status}: {txt[:160]}", status, retry_after
    except asyncio.TimeoutError:
        return None, "timeout", None, None
    except Exception as e:
        return None, str(e)[:100], None, None


async def call_with_key_pool(session, prompt, model, primary_provider,
                             key_pool, analytics, sub_id):
    """Send one prompt using the KEY POOL scheduler (replaces provider-level
    retries with key-level scheduling — requirement 10).

    Loop:
      1. Ask the pool for the next key (available first; else soonest-cooldown).
      2. If the chosen key is cooling down, sleep exactly until it's ready
         (this only happens when EVERY key is exhausted — requirement 6/7).
      3. Fire ONE request on that key.
         - 200  -> success, return.
         - 429  -> mark ONLY that key on cooldown, immediately try next key
                   (requirement 3/4). Never hammer the same key.
         - 5xx/timeout -> short cooldown on that key, try next.
         - 401/403 -> the key is invalid; it is marked dead and NEVER
           scheduled again (no retries on permanently bad keys).
         - other 4xx (bad request / unknown model) -> long cooldown; the key
           itself may be fine, so it isn't killed.
      4. Give up only after a bounded number of scheduling attempts with no
         success (avoids an infinite loop if all keys are permanently bad).

    The `model` is used when the chosen key's provider == primary_provider;
    otherwise the provider's own default model is used (so a Groq fallback key
    doesn't get sent a Gemini model id).
    """
    if not key_pool.has_any():
        return None, "no API keys configured"

    est_in = est_tokens(prompt)
    # Bound total scheduling attempts: enough to let every key try a few times.
    max_attempts = max(MAX_RETRIES + 1, len(key_pool.keys) * (MAX_RETRIES + 1))
    tried_notes = []

    for _attempt in range(max_attempts):
        now = time.time()
        api_key, wait = key_pool.acquire(now)
        if api_key is None:
            if key_pool.has_any():
                return None, "all API keys are invalid (authentication failed)"
            return None, "no API keys configured"

        # Every key is cooling down -> wait for the soonest (requirement 6/7).
        if wait > 0:
            await asyncio.sleep(min(wait, BACKOFF_CAP))
            now = time.time()

        cfg = SUPPORTED_APIS.get(api_key.provider)
        if cfg is None:
            api_key.mark_other_failure(now)
            continue

        # model: primary provider uses the user-picked model; others use default
        use_model = model if api_key.provider == primary_provider else cfg["models"][0]
        is_fallback = api_key.provider != primary_provider

        api_key.total_requests += 1
        api_key.last_request_time = now

        t0 = time.time()
        content, err, status, retry_after = await _single_call(
            session, cfg, api_key.key, use_model, prompt)
        latency = time.time() - t0

        analytics.record(
            provider=api_key.provider,
            key_label=api_key.label,
            model=use_model,
            request_no=analytics.next_request_no(),
            chars_sent=len(prompt),
            est_input_tokens=est_in,
            est_output_tokens=est_tokens(content) if content else 0,
            latency=round(latency, 2),
            retries=api_key.retry_count,
            status=status,
            fallback_used=is_fallback,
            submittal=sub_id,
        )

        if content is not None:
            api_key.mark_success()
            return content, None

        # Failure handling — per key.
        now = time.time()
        if status == 429:
            api_key.mark_429(now, retry_after)  # cooldown THIS key only,
            tried_notes.append(f"{api_key.label}:429")
            continue                            # immediately try next key
        elif status in RETRYABLE_STATUS or status is None:
            api_key.mark_other_failure(now)     # 5xx / timeout
            tried_notes.append(f"{api_key.label}:{status or 'net'}")
            continue
        elif status in (401, 403):
            # Invalid/revoked key: remove it from rotation permanently.
            api_key.mark_dead()
            tried_notes.append(f"{api_key.label}:{status}-dead")
            if not key_pool.has_usable():
                return None, "all API keys are invalid (authentication failed)"
            continue
        else:
            # Other 4xx (bad request / unknown model): the KEY may be fine, so
            # park it on a long cooldown instead of killing it, and move on.
            api_key.cooldown_until = now + 120.0
            api_key.total_failures += 1
            tried_notes.append(f"{api_key.label}:{status}")
            continue

    return None, f"all keys failed ({'; '.join(tried_notes[:8])})"


def _parse_comments(content):
    """Robustly parse the LLM's JSON into a list of comment strings.
    Handles markdown fences, leading prose, and malformed-but-salvageable JSON."""
    raw_content = content
    content = content.strip()
    content = re.sub(r'^```(?:json)?\s*', '', content)
    content = re.sub(r'\s*```$', '', content)

    parsed = None
    try:
        start = content.find("{")
        end = content.rfind("}")
        if start != -1 and end != -1 and end > start:
            parsed = json.loads(content[start:end + 1])
    except Exception:
        parsed = None

    if parsed is None:
        # Salvage a comments array even if the outer JSON is malformed.
        try:
            m = re.search(r'"comments"\s*:\s*\[(.*?)\]', content, re.S)
            if m:
                items = re.findall(r'"((?:[^"\\]|\\.)*)"', m.group(1))
                def _unescape(s):
                    # JSON-decode the escapes; unlike unicode_escape this
                    # never mangles non-ASCII characters.
                    try:
                        return json.loads('"' + s + '"')
                    except Exception:
                        return s
                parsed = {"comments": [_unescape(i) for i in items]}
        except Exception:
            parsed = None

    if parsed is None:
        return None, raw_content

    return parsed.get("comments", []), None


def _clean_comment_list(raw):
    """Apply the existing cleaning pipeline (order preserved)."""
    cleaned = []
    for c in raw:
        if not isinstance(c, str):
            continue
        c = re.sub(r'^\s*\d+\s*[\.\)]\s*', '', c).strip()
        if c:
            cleaned.append(c)
    cleaned = split_merged_comments(cleaned)
    cleaned = drop_headers(cleaned)
    cleaned = [strip_leadin(c) for c in cleaned]
    cleaned = [c for c in cleaned if len(c) >= 3]
    return cleaned


async def _llm_only(session, file_name, pdf_text, doc_name, unreadable,
                    api_choice, model, key_pool, analytics, trace=None):
    """Extract comments for one submittal (ZIP or standalone PDF). Splits
    oversized text, calls the key-pool scheduler per chunk, merges results
    in order. `trace` (Debug Mode) records prompts and raw responses."""
    sub_id = submittal_id_from_name(file_name)
    base = {"submittal": sub_id, "document": doc_name, "unreadable": unreadable}
    if not pdf_text.strip():
        return {**base, "comments": [], "error": "No text in PDF"}

    # Nothing is dropped: an oversized document becomes MULTIPLE requests
    # (each under the per-request budget) whose results are merged. Only a
    # pathological document beyond MAX_CHUNKS_PER_DOC ever loses content.
    chunks = split_oversized(pdf_text, min(CHAR_BUDGET, MODEL_CHAR_LIMIT))
    chunks = cap_chunks(chunks)
    if trace is not None:
        trace.chunks = len(chunks)

    all_comments = []
    last_err = None
    for ci, chunk in enumerate(chunks, 1):
        prompt = PROMPT_TEMPLATE.format(body=chunk)
        if trace is not None:
            trace.prompts.append(prompt)
        content, err = await call_with_key_pool(
            session, prompt, model, api_choice, key_pool, analytics, sub_id
        )
        if trace is not None:
            trace.responses.append({"chunk": ci, "content": content,
                                    "error": err, "parse_error": None})
        if err or content is None:
            last_err = err or "no content"
            print(f"DEBUG {sub_id}: {last_err}")
            continue
        parsed, _raw = _parse_comments(content)
        if parsed is None:
            last_err = f"No JSON in response (got: {(_raw or '')[:40]})"
            if trace is not None:
                trace.responses[-1]["parse_error"] = last_err
            print(f"DEBUG {sub_id}: {last_err}")
            continue
        all_comments.extend(_clean_comment_list(parsed))

    # De-duplicate across chunks/pages while preserving order.
    all_comments = dedupe_comments(all_comments)

    # If we got comments from at least one chunk, that's a success even if
    # another chunk failed (partial recovery beats total failure).
    if all_comments:
        return {**base, "comments": all_comments, "error": None}
    return {**base, "comments": [], "error": last_err or "no comments extracted"}


async def _process_one(session, file_name, file_bytes, sem,
                       api_choice, model, key_pool, analytics, debug_dir=None):
    """Extract text for ONE upload (ZIP or standalone PDF) and call the LLM,
    under the semaphore so only CONCURRENCY files are parsed at once. Frees
    PDF bytes and gc's between. One file failing never stops the batch
    (exception is caught -> error row). `debug_dir` (Debug Mode) writes the
    full pipeline trace to debug_dir/<SUBMITTAL>/ — diagnostics only."""
    async with sem:
        trace = DebugTrace(submittal_id_from_name(file_name)) \
            if debug_dir else None
        try:
            # A directly-uploaded PDF has no sibling files, so it is always
            # treated as the comment-bearing document itself.
            is_single_pdf = file_name.lower().endswith('.pdf')
            pdf_items = _pdf_items_from_upload(file_name, file_bytes)
            file_bytes = None
            pages, page_annots, unreadable, had_comment_file = \
                extract_pdf_text(pdf_items, treat_all_as_comments=is_single_pdf,
                                 trace=trace)
            doc_name = extract_doc_name_from_items(pdf_items)
            # Vendor text of the unannotated base twin (if the upload has
            # one): used AFTER the LLM step to drop comments that are really
            # the vendor's own drawing notes, not reviewer additions.
            vendor_blob = "" if is_single_pdf else _vendor_reference_text(pdf_items)
            pdf_items = None
            if not had_comment_file:
                # No annotated/response PDF -> genuinely no reviewer comments.
                res = {"submittal": submittal_id_from_name(file_name),
                       "document": doc_name, "unreadable": 0,
                       "comments": [], "error": None}
            else:
                # Token reduction: strip repeated boilerplate, then re-attach
                # each page's annotation text (page order preserved;
                # annotations are never subject to boilerplate removal).
                pages = dedupe_boilerplate(pages)
                blocks = []
                for i, ptxt in enumerate(pages):
                    ann = page_annots[i] if i < len(page_annots) else []
                    if ann:
                        ptxt = (ptxt + "\n" + "\n".join(
                            f"[Reviewer annotation] {t}" for t in ann)).strip()
                    if ptxt.strip():
                        blocks.append(ptxt)
                text = "\n".join(blocks)
                pages = page_annots = None
                gc.collect()
                res = await _llm_only(session, file_name, text, doc_name,
                                      unreadable, api_choice, model,
                                      key_pool, analytics, trace=trace)
                # Deterministic vendor filter: drop extracted "comments" that
                # are textually contained in the unannotated base twin —
                # vendor drawing notes, never reviewer additions. The model
                # input was untouched, so genuine-comment recall is exactly
                # what the model already delivered.
                if vendor_blob and res.get("comments"):
                    kept = []
                    for c in res["comments"]:
                        if _is_vendor_comment(c, vendor_blob):
                            if trace is not None:
                                trace.notes.append(
                                    f"vendor comment removed (matches base "
                                    f"twin): {c[:90]}")
                        else:
                            kept.append(c)
                    res["comments"] = kept
        except Exception as e:
            if trace is not None:
                trace.notes.append(f"processing exception: {str(e)[:80]}")
            res = {"submittal": submittal_id_from_name(file_name),
                   "document": "", "unreadable": 0,
                   "comments": [], "error": f"Processing error: {str(e)[:60]}"}
        if trace is not None:
            # Debug Mode must never break processing: trace write failures are
            # reported to the console and otherwise ignored.
            try:
                trace.save(debug_dir, res, analytics)
            except Exception as e:
                print(f"DEBUG-MODE: could not write trace for "
                      f"{trace.sub_id}: {e}")
        gc.collect()
        return res


async def process_all(zip_files, api_choice, model,
                      key_pool, analytics, progress_cb=None, debug_dir=None):
    sem = asyncio.Semaphore(CONCURRENCY)
    results = []
    async with aiohttp.ClientSession() as session:
        tasks = [
            _process_one(session, file_name, file_bytes, sem,
                         api_choice, model, key_pool, analytics,
                         debug_dir=debug_dir)
            for file_name, file_bytes in zip_files
        ]
        done = 0
        for coro in asyncio.as_completed(tasks):
            res = await coro
            results.append(res)
            done += 1
            if progress_cb:
                progress_cb(done, len(tasks))
    order = {submittal_id_from_name(n): i for i, (n, _) in enumerate(zip_files)}
    results.sort(key=lambda r: order.get(r["submittal"], 999))
    return results


# ============================================================================
# EXCEL (matches TPL_Comments.xlsx exactly)  — unchanged
# ============================================================================
def build_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font  = Font(bold=True, size=11)
    header_fill  = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cnr_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = ["Sr no.", "Submittal", "Document ", "Costumer Comments",
               "Xylem Remarks", "Review Note"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = center; c.border = border

    row = 2
    sr = 1
    for r in results:
        comments = r["comments"]
        has = bool(comments)
        n = max(1, len(comments))

        for k in range(n):
            first = (k == 0)
            c = ws.cell(row=row, column=1, value=(sr if first else None))
            c.alignment = center; c.border = border
            c = ws.cell(row=row, column=2, value=(r["submittal"] if first else None))
            c.alignment = left; c.border = border
            c = ws.cell(row=row, column=3, value=(r.get("document", "") if first else None))
            c.alignment = left; c.border = border
            val = f"{k+1}. {comments[k]}" if has else ""
            c = ws.cell(row=row, column=4, value=val)
            c.alignment = left; c.border = border
            if first and not has:
                c = ws.cell(row=row, column=5, value="Comment not Received")
                c.fill = cnr_fill
            else:
                c = ws.cell(row=row, column=5, value=None)
            c.alignment = left; c.border = border
            if first:
                unread = r.get("unreadable", 0)
                if unread > 0:
                    note = (f"{unread} page(s) could not be read as text "
                            f"(possible scanned comments) — check PDF manually")
                else:
                    note = ""
                cn = ws.cell(row=row, column=6, value=note)
                if unread > 0:
                    cn.fill = cnr_fill
            else:
                cn = ws.cell(row=row, column=6, value=None)
            cn.alignment = left; cn.border = border
            row += 1

        sr += 1

    last = row - 1
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 69
    ws.column_dimensions["E"].width = 21
    ws.column_dimensions["F"].width = 30
    ws.freeze_panes = "A2"
    if last >= 1:
        ws.auto_filter.ref = f"A1:F{last}"
    return wb


# ============================================================================
# UI  (unchanged layout; adds optional fallback keys + analytics summary)
# ============================================================================
st.title("TPL Comment Extractor")
st.caption("Upload submittal ZIPs or standalone PDFs. Comments are extracted "
           "by an LLM in parallel, split into individual items, and written "
           "to the TPL Excel format.")

# API selection — in the main page so it's always visible (mobile-friendly)
st.subheader("1. Choose API")
col1, col2 = st.columns(2)
with col1:
    api_choice = st.selectbox("Primary API Service", list(SUPPORTED_APIS.keys()),
                              help="Which provider to prefer first. Its keys are "
                                   "used before falling back to other providers.")
with col2:
    available_models = SUPPORTED_APIS[api_choice]["models"]
    model = st.selectbox("Model", available_models,
                         help="Pick a model, or type a custom one below")

custom_model = st.text_input("Custom model (optional)", value="",
                             help="Override the model above with any model ID")
if custom_model.strip():
    model = custom_model.strip()

# ---------------------------------------------------------------------------
# PROVIDER POOL UI — add/remove multiple keys per provider. Every provider in
# SUPPORTED_APIS automatically gets its own key section (nothing hard-coded).
# Keys live in st.session_state so the +/- buttons persist across reruns.
# Backward compatible: enter one Gemini + one Groq key and it behaves as before.
# ---------------------------------------------------------------------------
st.subheader("2. API Key Pool")
st.caption("Add multiple keys per provider. During processing the app rotates "
           "through them: a key that hits a rate limit (429) is put on cooldown "
           "and the next available key is used immediately. Scheduling priority: "
           + " → ".join(PROVIDER_PRIORITY)
           + ", then any other provider with keys.")

POOL_PROVIDERS = list(SUPPORTED_APIS.keys())

# initialise session state: one empty slot per provider on first load
for prov in POOL_PROVIDERS:
    sk = f"pool_{prov}"
    if sk not in st.session_state:
        st.session_state[sk] = [""]

def _add_key(prov):
    st.session_state[f"pool_{prov}"].append("")

def _remove_key(prov, idx):
    lst = st.session_state[f"pool_{prov}"]
    if 0 <= idx < len(lst):
        lst.pop(idx)
    if not lst:                       # never leave a provider with zero slots
        lst.append("")

for prov in POOL_PROVIDERS:
    sk = f"pool_{prov}"
    with st.expander(f"{prov.upper()} keys "
                     f"({sum(1 for k in st.session_state[sk] if k.strip())} set)",
                     expanded=(prov == api_choice)):
        for idx in range(len(st.session_state[sk])):
            kcol, bcol = st.columns([6, 1])
            with kcol:
                st.session_state[sk][idx] = st.text_input(
                    f"{prov.upper()} Key {idx + 1}",
                    value=st.session_state[sk][idx],
                    type="password",
                    key=f"{sk}_input_{idx}",
                )
            with bcol:
                st.write("")  # vertical spacer to align button with input
                st.button("➖", key=f"{sk}_del_{idx}",
                          help="Remove this key",
                          on_click=_remove_key, args=(prov, idx))
        st.button(f"➕ Add another {prov.upper()} key",
                  key=f"{sk}_add", on_click=_add_key, args=(prov,))

# Collect the configured pool (order preserved, blanks dropped).
provider_key_lists = {
    prov: [k.strip() for k in st.session_state[f"pool_{prov}"] if k.strip()]
    for prov in POOL_PROVIDERS
}
total_keys = sum(len(v) for v in provider_key_lists.values())

if st.button("Clear results & free memory"):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    gc.collect()
    st.rerun()

st.divider()
st.subheader("3. Upload & Extract")

if total_keys == 0:
    st.warning("Add at least one API key above (any provider) to begin.")
    st.stop()

# Make sure the primary provider actually has keys; if not, fall back to
# whichever provider does (so the user isn't forced to match the dropdown).
if not provider_key_lists.get(api_choice):
    wanted = api_choice
    for prov in POOL_PROVIDERS:
        if provider_key_lists.get(prov):
            api_choice = prov
            model = SUPPORTED_APIS[prov]["models"][0]
            st.info(f"No {wanted.upper()} keys entered for the selected "
                    f"primary; using {prov.upper()} as primary instead.")
            break

debug_mode = st.checkbox(
    "Enable Debug Mode",
    value=False,
    help="Diagnostics only — extraction behaviour is unchanged. Writes a full "
         "pipeline trace for every submittal to debug_logs/<SUBMITTAL>/: "
         "selected PDFs + reasons, per-page extracted text, annotations, the "
         "exact prompts, raw LLM responses, cleaned comments with source "
         "pages, Excel rows and a summary.")

uploaded = st.file_uploader("Upload ZIP or PDF files", type=["zip", "pdf"],
                            accept_multiple_files=True)

if uploaded:
    st.info(f"{len(uploaded)} file(s) ready. {total_keys} key(s) in pool.")
    if st.button("Extract Comments"):
        zip_files = [(f.name, f.read()) for f in uploaded]

        # Build the key pool from every configured key.
        key_pool = KeyPool(provider_key_lists)

        analytics = Analytics()
        wall_start = time.time()

        bar = st.progress(0.0)
        status = st.empty()
        def cb(done, total):
            bar.progress(done / total)
            status.text(f"Processed {done}/{total} submittals...")

        status.text(f"Sending to {api_choice.upper()} (pool of {total_keys} keys)...")
        results = asyncio.run(process_all(
            zip_files, api_choice, model, key_pool, analytics, cb,
            debug_dir=(DEBUG_DIR if debug_mode else None)))
        status.text("Building Excel...")

        wb = build_excel(results)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            data = open(tmp.name, "rb").read()
        os.unlink(tmp.name)

        wall_seconds = time.time() - wall_start

        st.success("Done.")
        if debug_mode:
            st.info(f"Debug Mode: pipeline traces written to "
                    f"`{os.path.abspath(DEBUG_DIR)}` "
                    f"(one folder per submittal).")
        st.download_button(
            "Download TPL_Comments.xlsx",
            data=data,
            file_name=f"TPL_Comments_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        total = sum(len(r["comments"]) for r in results)
        errs = [r for r in results if r["error"]]
        ok_zips = len(results) - len(errs)

        st.subheader("Summary")
        st.write(f"Submittals processed: {len(results)}")
        st.write(f"Total comments extracted: {total}")
        st.write(f"Primary API: {api_choice.upper()}  |  Keys in pool: {total_keys}")
        if errs:
            st.warning(f"{len(errs)} submittal(s) had issues:")
            for r in errs:
                st.write(f"- {r['submittal']}: {r['error']}")

        # ---- Request analytics summary ------------------------------------
        summary = analytics.summary(
            total_zips=len(results),
            ok_zips=ok_zips,
            fail_zips=len(errs),
            wall_seconds=wall_seconds,
        )
        st.subheader("Request Analytics")
        st.table(summary)

        # ---- Per-key analytics (requirement 9) ----------------------------
        st.subheader("Per-Key Statistics")
        st.table(analytics.per_key_table(key_pool))

        with st.expander("Preview extracted comments"):
            for r in results:
                st.markdown(f"**{r['submittal']}** — {len(r['comments'])} comment(s)")
                for i, c in enumerate(r["comments"], 1):
                    st.write(f"{i}. {c}")

        gc.collect()
